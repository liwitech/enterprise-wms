import io
from datetime import datetime
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_db, require_role
from app.core.security import get_password_hash
from app.models.department import Department
from app.models.enums import UserRoleEnum
from app.models.organization import Organization
from app.models.user import User
from app.schemas.admin import AdminDeptCreate, AdminDeptUpdate, AdminUserCreate, AdminUserUpdate
from app.schemas.common import ApiResponse, ok, paginated
from app.schemas.department import DepartmentRead, DepartmentTree
from app.schemas.organization import OrganizationRead, OrganizationUpdate
from app.schemas.user import UserRead

router = APIRouter(prefix="/admin", tags=["admin"])

_ADMIN = require_role(UserRoleEnum.SUPER_ADMIN, UserRoleEnum.ADMIN)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _build_tree(depts: list, parent_id=None) -> list[DepartmentTree]:
    nodes = []
    for d in depts:
        if d.parent_dept_id == parent_id:
            nodes.append(
                DepartmentTree(
                    id=d.id,
                    org_id=d.org_id,
                    name=d.name,
                    code=d.code,
                    dept_type=d.dept_type,
                    parent_dept_id=d.parent_dept_id,
                    manager_user_id=d.manager_user_id,
                    created_at=d.created_at,
                    children=_build_tree(depts, d.id),
                )
            )
    return sorted(nodes, key=lambda x: x.name)


# ── Organization ──────────────────────────────────────────────────────────────

@router.get("/organization", response_model=ApiResponse[OrganizationRead])
async def get_organization(
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    org = await db.get(Organization, current_user.org_id)
    if not org:
        raise HTTPException(404, "Organization not found")
    return ok(OrganizationRead.model_validate(org))


@router.put("/organization", response_model=ApiResponse[OrganizationRead])
async def update_organization(
    body: OrganizationUpdate,
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    org = await db.get(Organization, current_user.org_id)
    if not org:
        raise HTTPException(404, "Organization not found")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(org, field, value)
    await db.commit()
    await db.refresh(org)
    return ok(OrganizationRead.model_validate(org))


# ── Departments ───────────────────────────────────────────────────────────────

@router.get("/departments", response_model=ApiResponse[list[DepartmentTree]])
async def list_departments_tree(
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(
        select(Department)
        .where(Department.org_id == current_user.org_id)
        .order_by(Department.name)
    )).scalars().all()
    return ok(_build_tree(list(rows)))


@router.get("/departments/flat", response_model=ApiResponse[list[DepartmentRead]])
async def list_departments_flat(
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(
        select(Department)
        .where(Department.org_id == current_user.org_id)
        .order_by(Department.name)
    )).scalars().all()
    return ok([DepartmentRead.model_validate(d) for d in rows])


@router.post("/departments", response_model=ApiResponse[DepartmentRead], status_code=201)
async def create_department(
    body: AdminDeptCreate,
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    dept = Department(
        org_id=current_user.org_id,
        name=body.name,
        code=body.code,
        dept_type=body.dept_type.value if body.dept_type else None,
        parent_dept_id=body.parent_dept_id,
        manager_user_id=body.manager_user_id,
    )
    db.add(dept)
    await db.commit()
    await db.refresh(dept)
    return ok(DepartmentRead.model_validate(dept))


@router.put("/departments/{dept_id}", response_model=ApiResponse[DepartmentRead])
async def update_department(
    dept_id: UUID,
    body: AdminDeptUpdate,
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    dept = await db.get(Department, dept_id)
    if not dept or dept.org_id != current_user.org_id:
        raise HTTPException(404, "Department not found")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(dept, field, value)
    await db.commit()
    await db.refresh(dept)
    return ok(DepartmentRead.model_validate(dept))


@router.delete("/departments/{dept_id}", response_model=ApiResponse[None])
async def delete_department(
    dept_id: UUID,
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    dept = await db.get(Department, dept_id)
    if not dept or dept.org_id != current_user.org_id:
        raise HTTPException(404, "Department not found")
    child_count = (await db.execute(
        select(func.count()).select_from(Department).where(Department.parent_dept_id == dept_id)
    )).scalar_one()
    if child_count > 0:
        raise HTTPException(400, "Không thể xóa đơn vị đang có đơn vị con")
    await db.delete(dept)
    await db.commit()
    return ok(None)


# ── Users / Employees ─────────────────────────────────────────────────────────

@router.get("/users", response_model=ApiResponse[list[UserRead]])
async def list_users(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: str | None = None,
    dept_id: UUID | None = None,
    role: UserRoleEnum | None = None,
    is_active: bool | None = None,
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    q = select(User).where(
        User.org_id == current_user.org_id,
        User.deleted_at.is_(None),
    )
    if search:
        q = q.where(
            or_(
                User.full_name.ilike(f"%{search}%"),
                User.email.ilike(f"%{search}%"),
                User.employee_code.ilike(f"%{search}%"),
            )
        )
    if dept_id:
        q = q.where(User.dept_id == dept_id)
    if role:
        q = q.where(User.role == role)
    if is_active is not None:
        q = q.where(User.is_active == is_active)

    total = (await db.execute(
        select(func.count()).select_from(q.subquery())
    )).scalar_one()
    rows = (await db.execute(
        q.order_by(User.full_name).offset((page - 1) * per_page).limit(per_page)
    )).scalars().all()

    return paginated([UserRead.model_validate(u) for u in rows], page, per_page, total)


@router.post("/users", response_model=ApiResponse[UserRead], status_code=201)
async def create_user(
    body: AdminUserCreate,
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    existing = (await db.execute(
        select(User).where(User.email == body.email, User.deleted_at.is_(None))
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(400, "Email đã được sử dụng")

    user = User(
        org_id=current_user.org_id,
        dept_id=body.dept_id,
        email=body.email,
        full_name=body.full_name,
        hashed_password=get_password_hash(body.password),
        role=body.role,
        employee_code=body.employee_code,
        is_active=True,
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return ok(UserRead.model_validate(user))


@router.get("/users/{user_id}", response_model=ApiResponse[UserRead])
async def get_user(
    user_id: UUID,
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    user = await db.get(User, user_id)
    if not user or user.org_id != current_user.org_id or user.deleted_at is not None:
        raise HTTPException(404, "User not found")
    return ok(UserRead.model_validate(user))


@router.put("/users/{user_id}", response_model=ApiResponse[UserRead])
async def update_user(
    user_id: UUID,
    body: AdminUserUpdate,
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    user = await db.get(User, user_id)
    if not user or user.org_id != current_user.org_id or user.deleted_at is not None:
        raise HTTPException(404, "User not found")

    data = body.model_dump(exclude_none=True)
    if "password" in data:
        user.hashed_password = get_password_hash(data.pop("password"))
    for field, value in data.items():
        setattr(user, field, value)

    await db.commit()
    await db.refresh(user)
    return ok(UserRead.model_validate(user))


@router.patch("/users/{user_id}/toggle-active", response_model=ApiResponse[UserRead])
async def toggle_user_active(
    user_id: UUID,
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    user = await db.get(User, user_id)
    if not user or user.org_id != current_user.org_id or user.deleted_at is not None:
        raise HTTPException(404, "User not found")
    user.is_active = not user.is_active
    await db.commit()
    await db.refresh(user)
    return ok(UserRead.model_validate(user))


@router.delete("/users/{user_id}", response_model=ApiResponse[None])
async def delete_user(
    user_id: UUID,
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    user = await db.get(User, user_id)
    if not user or user.org_id != current_user.org_id or user.deleted_at is not None:
        raise HTTPException(404, "User not found")
    if user.id == current_user.id:
        raise HTTPException(400, "Không thể xóa tài khoản đang đăng nhập")
    user.deleted_at = datetime.utcnow()
    user.is_active = False
    await db.commit()
    return ok(None)


# ── User Import ───────────────────────────────────────────────────────────────

_ROLE_MAP = {
    "SUPER_ADMIN": UserRoleEnum.SUPER_ADMIN,
    "ADMIN": UserRoleEnum.ADMIN,
    "MANAGER": UserRoleEnum.MANAGER,
    "EMPLOYEE": UserRoleEnum.EMPLOYEE,
    "Siêu quản trị": UserRoleEnum.SUPER_ADMIN,
    "Quản trị viên": UserRoleEnum.ADMIN,
    "Quản lý": UserRoleEnum.MANAGER,
    "Nhân viên": UserRoleEnum.EMPLOYEE,
}

_TEMPLATE_HEADERS = [
    "Họ và tên *",
    "Email *",
    "Mật khẩu *",
    "Mã nhân viên",
    "Vai trò",
    "Mã phòng ban",
]

_TEMPLATE_NOTES = [
    "Bắt buộc. Họ tên đầy đủ.",
    "Bắt buộc. Phải là email hợp lệ và chưa tồn tại.",
    "Bắt buộc. Tối thiểu 6 ký tự.",
    "Tùy chọn. Mã định danh nhân viên (duy nhất).",
    "Tùy chọn. Một trong: SUPER_ADMIN, ADMIN, MANAGER, EMPLOYEE. Mặc định: EMPLOYEE.",
    "Tùy chọn. Mã đơn vị (cột 'code' trong bảng phòng ban).",
]

_TEMPLATE_SAMPLE = [
    "Nguyễn Văn An",
    "an.nguyen@company.vn",
    "Password@123",
    "NV001",
    "EMPLOYEE",
    "PHONG_KD",
]


@router.get("/users/import/template")
async def download_import_template(
    current_user: User = Depends(_ADMIN),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách người dùng"

    # ── Header row
    header_fill = PatternFill("solid", fgColor="C0392B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(_TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 30

    # ── Sample row
    sample_fill = PatternFill("solid", fgColor="FFF9F9")
    for col, val in enumerate(_TEMPLATE_SAMPLE, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = sample_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    # ── Notes row
    note_fill = PatternFill("solid", fgColor="F5F5F5")
    note_font = Font(italic=True, color="888888", size=9)
    for col, note in enumerate(_TEMPLATE_NOTES, 1):
        cell = ws.cell(row=3, column=col, value=note)
        cell.fill = note_fill
        cell.font = note_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[3].height = 45

    # ── Column widths
    col_widths = [25, 30, 18, 16, 18, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze header
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="import_users_template.xlsx"'},
    )


class ImportErrorRow(BaseModel):
    row: int
    full_name: str
    email: str
    message: str


class ImportResult(BaseModel):
    total: int
    success: int
    failed: int
    errors: list[ImportErrorRow]


@router.post("/users/import", response_model=ApiResponse[ImportResult])
async def import_users(
    file: UploadFile = File(...),
    current_user: User = Depends(_ADMIN),
    db: AsyncSession = Depends(get_db),
):
    from openpyxl import load_workbook

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Chỉ hỗ trợ file Excel (.xlsx, .xls)")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "File không hợp lệ hoặc bị lỗi")

    # Lấy tất cả departments trong org để tra cứu theo code
    depts = (await db.execute(
        select(Department).where(Department.org_id == current_user.org_id)
    )).scalars().all()
    dept_by_code = {d.code.upper(): d.id for d in depts}

    # Lấy tất cả email và employee_code đang dùng
    existing_emails = set(
        r[0] for r in (await db.execute(
            select(User.email).where(User.org_id == current_user.org_id, User.deleted_at.is_(None))
        )).all()
    )
    existing_codes = set(
        r[0] for r in (await db.execute(
            select(User.employee_code).where(
                User.org_id == current_user.org_id,
                User.deleted_at.is_(None),
                User.employee_code.isnot(None),
            )
        )).all()
    )

    errors: list[ImportErrorRow] = []
    success = 0
    # Bỏ qua dòng 1 (header), dòng 2 (sample), dòng 3 (notes) nếu chứa text mô tả
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Phát hiện dòng ghi chú — bỏ qua nếu cột email không có @
    data_rows = [
        (excel_row, r) for excel_row, r in enumerate(rows, 2)
        if r and any(r) and "@" in str(r[1] or "")
    ]

    for excel_row, row in data_rows:
        full_name = str(row[0] or "").strip()
        email = str(row[1] or "").strip().lower()
        password = str(row[2] or "").strip()
        employee_code = str(row[3] or "").strip() or None
        role_raw = str(row[4] or "").strip().upper() or "EMPLOYEE"
        dept_code = str(row[5] or "").strip().upper() or None

        # Validate
        issues = []
        if not full_name:
            issues.append("Thiếu họ và tên")
        if not email or "@" not in email:
            issues.append("Email không hợp lệ")
        elif email in existing_emails:
            issues.append(f"Email '{email}' đã tồn tại")
        if not password or len(password) < 6:
            issues.append("Mật khẩu tối thiểu 6 ký tự")
        if employee_code and employee_code in existing_codes:
            issues.append(f"Mã nhân viên '{employee_code}' đã tồn tại")

        role = _ROLE_MAP.get(role_raw)
        if not role:
            issues.append(f"Vai trò '{role_raw}' không hợp lệ")

        dept_id = None
        if dept_code:
            dept_id = dept_by_code.get(dept_code)
            if not dept_id:
                issues.append(f"Mã phòng ban '{dept_code}' không tồn tại")

        if issues:
            errors.append(ImportErrorRow(
                row=excel_row,
                full_name=full_name or "(trống)",
                email=email or "(trống)",
                message="; ".join(issues),
            ))
            continue

        # Tạo user
        user = User(
            org_id=current_user.org_id,
            dept_id=dept_id,
            email=email,
            full_name=full_name,
            hashed_password=get_password_hash(password),
            role=role,
            employee_code=employee_code,
            is_active=True,
        )
        db.add(user)
        existing_emails.add(email)
        if employee_code:
            existing_codes.add(employee_code)
        success += 1

    await db.commit()

    total = success + len(errors)
    return ok(ImportResult(total=total, success=success, failed=len(errors), errors=errors))
